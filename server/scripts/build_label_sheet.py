"""Build the hand-labelling sheet: the first 50 calls from the dashboard, in English.

WHY THIS EXISTS
CallHarness's verdicts have never been checked against a human. Until they are, no
number the dashboard shows can be trusted, and no agent fix can be measured. This
script produces the answer key: 50 real calls, translated, with their tool calls laid
out, ready for a human to label by hand. Once labelled, the same 50 become a permanent
regression test — every future change to the judge gets scored against them in seconds.

The sheet HIDES CallHarness's own verdict (hidden columns, not absent) so the labeller
is not anchored by it. The Score tab reads those columns and does the comparison.

USAGE
    cd server
    .venv\\Scripts\\activate
    pip install openpyxl httpx
    python -m scripts.build_label_sheet --out ..\\data\\label_sheet_50.xlsx

Translation uses the server's existing POST /calls/{id}/translate endpoint, so the
English is cached onto the call and shows in the dashboard too — you pay once.
Re-runs are free. Pass --no-translate to skip entirely.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from typing import Any

try:
    import httpx
except ImportError:  # pragma: no cover
    sys.exit("Missing dependency. Run:  pip install httpx openpyxl")

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover
    sys.exit("Missing dependency. Run:  pip install openpyxl httpx")


# The six fault codes. Six and only six: a taxonomy you can hold in your head is
# one you apply consistently. Each came out of calls already seen in the Lazio data.
FAULTS: list[tuple[str, str]] = [
    ("missing_data",
     "Agent DID look it up, lookup came back empty. Record isn't in Cerba's database."),
    ("wrong_tool",
     "Agent called the wrong tool, or none at all. Data may exist — it never went for it."),
    ("stt",
     "Speech-to-text mangled the caller, so the agent searched for nonsense."),
    ("agent_hallucinated",
     "Lookup gave nothing, agent answered anyway. Invented hours/prices/availability."),
    ("caller_left",
     "Heard the greeting or a good answer and chose to go. Nobody failed."),
    ("no_caller_audio",
     "Caller never appears in the transcript at all. Don't guess why — needs audio."),
]
FAULT_KEYS = [f[0] for f in FAULTS]

# Result payloads meaning "the lookup ran and matched nothing". Mirrors
# app/knowledge_gaps.py::_EMPTY_MARKERS, copied so this script imports nothing.
EMPTY_MARKERS = (
    "no results", "no result", "not found", "no data", "no match", "empty",
    "nessun risultato", "non trovato", "non disponibile", "nessun dato",
    "non ho trovato", "non ho una risposta", "cerca nel rag",
    "informazioni specifiche",
)
TECHNICAL_MARKERS = (
    "timeout", "timed out", "connection", "refused", "unreachable",
    "502", "503", "504", "500 internal", "traceback", "exception",
    "ssl", "unauthorized", "forbidden", "rate limit",
)

FONT = "Arial"
NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow = yours
BAND_FILL = PatternFill("solid", fgColor="F7F9FC")     # zebra striping
FLAG_FILL = PatternFill("solid", fgColor="FBD9D3")
THIN = Side(style="thin", color="D0D7E5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _text_of(value: Any) -> str:
    if isinstance(value, str):
        return value.lower()
    try:
        return json.dumps(value, default=str, ensure_ascii=False).lower()
    except (TypeError, ValueError):
        return str(value).lower()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, (list, tuple, set)):
        return len(value) == 0
    if isinstance(value, dict):
        if not value:
            return True
        for key in ("results", "result", "data", "items", "answer", "records", "matches"):
            if key in value:
                return _is_blank(value[key])
    return False


def classify_result(result: Any) -> str:
    """MISS / ERROR / OK for one tool result.

    Deliberately ignores a `success: true` field — the graph returns success=true
    alongside "Non ho una risposta per questo cerca nel RAG", a miss wearing a flag.
    """
    if _is_blank(result):
        return "MISS"
    text = _text_of(result)
    if isinstance(result, dict) and result.get("error"):
        err = _text_of(result["error"])
        if any(m in err for m in TECHNICAL_MARKERS):
            return "ERROR"
        if any(m in err for m in EMPTY_MARKERS):
            return "MISS"
        return "ERROR"
    if any(m in text for m in TECHNICAL_MARKERS):
        return "ERROR"
    if any(m in text for m in EMPTY_MARKERS):
        return "MISS"
    return "OK"


def _short(value: Any, n: int) -> str:
    if value is None:
        return ""
    s = value if isinstance(value, str) else json.dumps(value, default=str, ensure_ascii=False)
    s = " ".join(s.split())
    return s if len(s) <= n else s[: n - 1] + "…"


def fetch(api: str, key: str | None, limit: int, agent: str | None,
          translate: bool) -> list[dict[str, Any]]:
    headers = {"X-API-Key": key} if key else {}
    params: dict[str, Any] = {"limit": limit}
    if agent:
        params["agent_id"] = agent

    with httpx.Client(base_url=api, timeout=120.0, headers=headers) as client:
        resp = client.get("/api/v1/calls", params=params)
        resp.raise_for_status()
        listing = resp.json()
        ids = [c["id"] for c in listing["items"]]
        print(f"{listing['total']} calls match; taking the first {len(ids)} "
              f"(newest first, same order as the dashboard)")

        calls: list[dict[str, Any]] = []
        for i, call_id in enumerate(ids, 1):
            if translate:
                try:
                    r = client.post(f"/api/v1/calls/{call_id}/translate",
                                    params={"language": "english"})
                    if r.status_code == 200:
                        calls.append(r.json())
                        print(f"  [{i:>2}/{len(ids)}] {call_id[:8]}  ok")
                        continue
                    print(f"  [{i:>2}/{len(ids)}] {call_id[:8]}  translate "
                          f"{r.status_code} — using original")
                except httpx.HTTPError as exc:
                    print(f"  [{i:>2}/{len(ids)}] {call_id[:8]}  translate failed "
                          f"({exc}) — using original")
            d = client.get(f"/api/v1/calls/{call_id}")
            d.raise_for_status()
            calls.append(d.json())
            if not translate:
                print(f"  [{i:>2}/{len(ids)}] {call_id[:8]}")
    return calls


def transcript_of(call: dict[str, Any]) -> tuple[str, bool, int]:
    """(readable English script with tool calls inline, caller_spoke, miss_count)."""
    lines: list[str] = []
    caller_spoke = False
    misses = 0
    for t in call.get("turns", []):
        speaker = "CALLER" if t["role"] == "user" else "agent "
        italian = " ".join((t.get("text") or "").split())
        english = " ".join((t.get("translated_text") or "").split()) or italian
        if t["role"] == "user" and english.strip():
            caller_spoke = True
        if english.strip():
            lines.append(f"{speaker}  {english}")
        for tc in (t.get("tool_calls") or []):
            verdict = classify_result(tc.get("result"))
            if verdict == "MISS":
                misses += 1
            lines.append(
                f"   [{verdict}] {tc.get('name')}({_short(tc.get('arguments'), 90)})"
                f" → {_short(tc.get('result'), 200)}"
            )
    return "\n".join(lines), caller_spoke, misses


def build(calls: list[dict[str, Any]], dashboard: str, out_path: str) -> None:
    wb = Workbook()

    # ============================== THE TABLE ==============================
    ws = wb.active
    ws.title = "Label"

    cols = [
        ("#", 5),
        ("call", 11),
        ("secs", 6),
        ("spoke?", 8),
        ("miss", 6),
        ("transcript  (English · tool calls inline)", 78),
        ("what_caller_wanted", 26),
        ("got_it", 8),
        ("fault", 21),
        ("notes", 26),
        ("ch_outcome", 14),          # hidden ↓
        ("ch_success", 11),
        ("ch_transfer_reason", 20),
        ("ch_non_completion_reason", 24),
    ]

    # Row 1 — the only instructions there are.
    legend = (
        "Fill the YELLOW columns only.     "
        "got_it = did the caller leave with what they called for?  Y / N     "
        "fault (only when N) = missing_data · wrong_tool · stt · "
        "agent_hallucinated · caller_left · no_caller_audio     "
        "Pick the FIRST thing that went wrong, not the last."
    )
    c = ws.cell(row=1, column=1, value=legend)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E5496")
    c.alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.row_dimensions[1].height = 24

    # Row 2 — headers.
    for i, (title, width) in enumerate(cols, start=1):
        c = ws.cell(row=2, column=i, value=title)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[2].height = 26

    ws.cell(row=2, column=8).comment = Comment(
        "Y or N.\n\nNot 'was the agent polite'. Not 'did the agent say something'.\n"
        "Did the caller GET what they called for.\n\n"
        "A transfer to a human is N — they did not get it from us.",
        "Cooky", height=160, width=320,
    )
    ws.cell(row=2, column=9).comment = Comment(
        "\n".join(f"{k} — {d}" for k, d in FAULTS),
        "Cooky", height=220, width=420,
    )
    ws.cell(row=2, column=6).comment = Comment(
        "[MISS] = the lookup ran and found nothing.\n"
        "[ERROR] = our infrastructure broke.\n"
        "[OK] = the lookup returned something usable.\n\n"
        "MISS is computed ignoring any success:true flag, because the graph "
        "returns success=true alongside 'Non ho una risposta'.",
        "Cooky", height=180, width=380,
    )

    for n, call in enumerate(calls, start=1):
        row = n + 2
        script, spoke, misses = transcript_of(call)
        band = BAND_FILL if n % 2 == 0 else None

        values = [
            n,
            None,                                   # hyperlink below
            round(call.get("duration_seconds") or 0),
            "yes" if spoke else "NO",
            misses or "",
            script or "— no turns recorded —",
            None, None, None, None,                 # yours
            call.get("outcome"),
            call.get("success"),
            call.get("transfer_reason"),
            call.get("non_completion_reason"),
        ]
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.border = BORDER
            if i == 6:
                c.font = Font(name="Consolas", size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                c.font = Font(name=FONT, size=10)
                c.alignment = Alignment(
                    vertical="top",
                    horizontal="center" if i in (1, 3, 4, 5, 8) else "left",
                    wrap_text=(i in (7, 10)),
                )
            if 7 <= i <= 10:
                c.fill = INPUT_FILL
            elif band:
                c.fill = band

        c = ws.cell(row=row, column=2, value=call["id"][:8])
        c.hyperlink = f"{dashboard.rstrip('/')}/calls/{call['id']}"
        c.font = Font(name=FONT, size=10, color="0563C1", underline="single")
        c.alignment = Alignment(vertical="top", horizontal="center")
        c.border = BORDER
        if band:
            c.fill = band

        if not spoke:
            ws.cell(row=row, column=4).fill = FLAG_FILL
        if misses:
            ws.cell(row=row, column=5).fill = FLAG_FILL

        ws.row_dimensions[row].height = 150

    last = len(calls) + 2

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True,
                        showDropDown=False)
    yn.error = "Y or N only."
    ws.add_data_validation(yn)
    yn.add(f"H3:H{last}")

    fault_dv = DataValidation(type="list", formula1=f'"{",".join(FAULT_KEYS)}"',
                              allow_blank=True, showDropDown=False)
    fault_dv.error = "One of the six codes. Hover the header for definitions."
    ws.add_data_validation(fault_dv)
    fault_dv.add(f"I3:I{last}")

    for col in ("K", "L", "M", "N"):
        ws.column_dimensions[col].hidden = True

    # Transcript stays on screen while you type in the yellow columns.
    ws.freeze_panes = "G3"
    ws.auto_filter.ref = f"A2:J{last}"
    ws.sheet_view.zoomScale = 100

    # ================================ SCORE ================================
    ws = wb.create_sheet("Score")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 66

    t = ws.cell(row=1, column=1,
                value="Don't open this until all 50 rows are filled")
    t.font = Font(name=FONT, size=13, bold=True, color="C00000")
    ws.merge_cells("A1:C1")

    def line(row: int, label: str, formula: str, comment: str,
             pct: bool = False, big: bool = False) -> None:
        a = ws.cell(row=row, column=1, value=label)
        a.font = Font(name=FONT, size=11 if big else 10, bold=True,
                      color="C00000" if big else "000000")
        b = ws.cell(row=row, column=2, value=formula)
        b.font = Font(name=FONT, size=11 if big else 10, bold=big)
        if pct:
            b.number_format = "0.0%"
        c = ws.cell(row=row, column=3, value=comment)
        c.font = Font(name=FONT, size=9, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30

    last = len(calls) + 2
    line(3, "rows you've labelled", f"=COUNTA(Label!H3:H{last})",
         f"Needs to hit {len(calls)} before anything below means anything.")
    line(4, "your success rate", f"=IFERROR(COUNTIF(Label!H3:H{last},\"Y\")"
                                 f"/COUNTA(Label!H3:H{last}),\"\")",
         "The honest number for this sample.", pct=True)
    line(5, "CallHarness success rate",
         f"=IFERROR(COUNTIF(Label!K3:K{last},\"completed\")"
         f"/COUNTA(Label!K3:K{last}),\"\")",
         "What the dashboard claims for the same calls.", pct=True)
    line(6, "OUTCOME AGREEMENT",
         f"=IFERROR((COUNTIFS(Label!H3:H{last},\"Y\",Label!K3:K{last},\"completed\")"
         f"+COUNTIFS(Label!H3:H{last},\"N\",Label!K3:K{last},\"<>completed\"))"
         f"/COUNTA(Label!H3:H{last}),\"\")",
         "THE number. How often CallHarness agrees with you. Under ~85% means fix "
         "the judge before you touch the agent — otherwise you can't tell whether "
         "a change helped.", pct=True, big=True)
    line(8, "false Completed  (CH yes, you no)",
         f"=COUNTIFS(Label!H3:H{last},\"N\",Label!K3:K{last},\"completed\")",
         "The dangerous direction — inflates what you report to Cerba.")
    line(9, "false Failed  (CH no, you yes)",
         f"=COUNTIFS(Label!H3:H{last},\"Y\",Label!K3:K{last},\"<>completed\")",
         "Understates you. Annoying, not dangerous.")

    ws.cell(row=11, column=1, value="Where the failures actually are").font = Font(
        name=FONT, size=12, bold=True, color=NAVY)
    for i, title in enumerate(("fault", "count"), start=1):
        c = ws.cell(row=12, column=i, value=title)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.border = BORDER
    for i, (key, desc) in enumerate(FAULTS):
        row = 13 + i
        ws.cell(row=row, column=1, value=key).font = Font(name=FONT, size=10)
        b = ws.cell(row=row, column=2, value=f"=COUNTIF(Label!I3:I{last},\"{key}\")")
        b.font = Font(name=FONT, size=10)
        c = ws.cell(row=row, column=3, value=desc)
        c.font = Font(name=FONT, size=9, color="595959")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for i2 in range(1, 3):
            ws.cell(row=row, column=i2).border = BORDER

    row = 13 + len(FAULTS)
    ws.cell(row=row, column=1, value="unlabelled failures").font = Font(
        name=FONT, size=10, italic=True, color="C00000")
    ws.cell(row=row, column=2,
            value=f"=COUNTIF(Label!H3:H{last},\"N\")-COUNTA(Label!I3:I{last})"
            ).font = Font(name=FONT, size=10)
    ws.cell(row=row, column=3, value="Should be 0 — every N needs a fault code."
            ).font = Font(name=FONT, size=9, color="595959")

    row += 2
    for text in (
        "Biggest count in the fault table is what you fix first. Not the most "
        "interesting one. Not the most technical one. The biggest one.",
        "If missing_data wins, most of the 14% is Cerba's problem and your job is "
        "the report, not the agent.",
        "Keep this file. Re-run the comparison after every change to the analysis "
        "prompt. Doing it once is the only reason it was worth doing at all.",
    ):
        c = ws.cell(row=row, column=1, value="•  " + text)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 30
        row += 1

    parent = os.path.dirname(os.path.abspath(out_path))
    os.makedirs(parent, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--api", default=os.environ.get("CALLHARNESS_API_URL",
                                                   "http://localhost:8010"))
    p.add_argument("--dashboard", default="http://localhost:3011")
    p.add_argument("--limit", type=int, default=50)
    p.add_argument("--agent", default=None, help="filter by agent_id (optional)")
    p.add_argument("--no-translate", action="store_true")
    p.add_argument("--out", default="label_sheet_50.xlsx")
    args = p.parse_args()

    key = os.environ.get("CALLHARNESS_API_KEY")
    calls = fetch(args.api, key, args.limit, args.agent, not args.no_translate)
    if not calls:
        sys.exit(f"No calls returned. Is the server running on {args.api}?")
    build(calls, args.dashboard, args.out)
    print(f"\nWrote {args.out} — {len(calls)} calls, one row each.")


if __name__ == "__main__":
    main()
